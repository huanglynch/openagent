# skills/excel_processor.py
# 全能 Excel 处理 Skill - 生产就绪终极版（v2.0）
# 完全兼容当前 MultiAgentSwarm 架构 + 安全铁律

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook


def tool_function(
        file_path: str,
        operation: str = "read",          # read / analyze / edit / batch_edit / write / create / pivot
        sheet_name: str = None,
        cell: str = None,                 # 单 cell edit 用
        value: any = None,
        data: dict = None,                # write/create 用
        output_filename: str = None,
        batch_edits: dict = None,         # {"A1": 100, "B2": "新标题"}
        pivot_config: dict = None         # {"index": "日期", "columns": "类别", "values": "金额", "aggfunc": "sum"}
) -> dict:
    """
    全能 Excel 处理器（仅限 uploads/ 目录）
    返回结构化数据 + Markdown 表格（WebUI 直接美观显示）
    """
    try:
        path = Path(file_path)
        upload_dir = Path("uploads")

        # 【安全铁律】强制 uploads/ 目录
        if not str(path.resolve()).startswith(str(upload_dir.resolve())):
            return {"success": False, "error": "安全铁律：仅允许处理 uploads/ 目录下的文件"}

        result = {"success": True, "operation": operation, "file_path": str(path)}

        # ====================== 读取 & 分析 ======================
        if operation in ["read", "analyze", "pivot"]:
            if not path.exists():
                return {"success": False, "error": f"文件不存在: {file_path}"}

            excel_file = pd.ExcelFile(path)
            result["sheets"] = excel_file.sheet_names

            target_sheet = sheet_name or excel_file.sheet_names[0]
            df = pd.read_excel(path, sheet_name=target_sheet)

            result["data"] = df.head(200).to_dict(orient="records")
            result["shape"] = [len(df), len(df.columns)]
            result["columns"] = list(df.columns)
            result["markdown"] = df.head(30).to_markdown(index=False)

            if operation == "analyze" or operation == "pivot":
                result["analysis"] = {
                    "row_count": len(df),
                    "col_count": len(df.columns),
                    "numeric_cols": list(df.select_dtypes(include="number").columns),
                    "sample": df.head(5).to_dict(orient="records")
                }

            # ====================== Pivot 透视表（最常用报表需求）======================
            if operation == "pivot" and pivot_config:
                try:
                    pivot = pd.pivot_table(df, **pivot_config)
                    result["pivot_markdown"] = pivot.to_markdown()
                    result["pivot_data"] = pivot.head(50).to_dict(orient="records")
                    result["analysis"]["pivot"] = f"已生成透视表（{len(pivot)} 行）"
                except Exception as e:
                    result["analysis"]["pivot_error"] = str(e)

            return result

        # ====================== 单单元格编辑 ======================
        elif operation == "edit":
            if not path.exists() or not cell or value is None:
                return {"success": False, "error": "edit 操作需要有效 cell 和 value"}
            wb = load_workbook(path)
            ws = wb[sheet_name or wb.sheetnames[0]]
            ws[cell] = value
            wb.save(path)
            result["message"] = f"单元格 {cell} 已更新为 {value}"
            return result

        # ====================== 批量编辑 ======================
        elif operation == "batch_edit" or (operation == "edit" and batch_edits):
            if not path.exists() or not batch_edits:
                return {"success": False, "error": "batch_edit 需要 batch_edits 参数"}
            wb = load_workbook(path)
            ws = wb[sheet_name or wb.sheetnames[0]]
            for c, v in batch_edits.items():
                ws[c] = v
            wb.save(path)
            result["message"] = f"批量更新 {len(batch_edits)} 个单元格完成"
            return result

        # ====================== 写入 / 创建新文件 ======================
        elif operation in ["write", "create"]:
            if not data:
                return {"success": False, "error": "write/create 需要 data 参数"}

            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = pd.DataFrame.from_dict(data, orient="index") if isinstance(data, dict) else pd.DataFrame(data)

            output_path = upload_dir / (output_filename or f"new_{path.name}" if operation == "create" else path.name)
            df.to_excel(output_path, index=False)

            result["saved_path"] = str(output_path)
            result["markdown"] = df.head(20).to_markdown(index=False)
            result["message"] = f"文件已保存：{output_path.name}"
            return result

        return {"success": False, "error": f"未知 operation: {operation}"}

    except Exception as e:
        return {"success": False, "error": str(e)}


# ====================== Tool Schema ======================
tool_schema = {
    "type": "function",
    "function": {
        "name": "excel_processor",
        "description": "全能 Excel 处理 Skill（仅限 uploads/）。支持 read/analyze/edit/batch_edit/write/create/pivot。返回 Markdown 表格 + 结构化数据。推荐优先用 analyze（带分析）或 pivot（透视表）。",
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Excel 文件路径（必须在 uploads/ 下）"},
                "operation": {
                    "type": "string",
                    "enum": ["read", "analyze", "edit", "batch_edit", "write", "create", "pivot"],
                    "default": "read"
                },
                "sheet_name": {"type": "string", "description": "指定工作表（可选，默认第一个）"},
                "cell": {"type": "string", "description": "单 cell edit：如 A1"},
                "value": {"type": "string", "description": "单 cell edit：新值"},
                "data": {"type": "object", "description": "write/create：数据（列表或 dict）"},
                "output_filename": {"type": "string", "description": "新文件名（可选）"},
                "batch_edits": {"type": "object", "description": "批量编辑：{\"A1\": 值, \"B2\": 值}"},
                "pivot_config": {"type": "object", "description": "透视表配置：{index, columns, values, aggfunc}"}
            },
            "required": ["file_path"]
        }
    }
}